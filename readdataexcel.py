import openpyxl

wb = openpyxl.load_workbook("C:\\DVF\\Scripting\\RestAPI\\student.xlsx")
ws = wb['Sheet1']
rows = ws.max_row
print(rows)

for i in range(1, rows+1):
    cells = ws.cell(row=i, column=i)
    print(cells.value)
